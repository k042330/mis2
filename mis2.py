from flask import Flask, request, render_template, send_file
from datetime import datetime
import pandas as pd
import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)


def calculate_time_differences(df):
    """
    计算从'接收'到下一个'发出'之间的时间差

    Args:
        df: 包含会话数据的DataFrame
    """
    # 确保数据中有必要的列
    if "发出/接收" not in df.columns or "会话时间" not in df.columns:
        raise ValueError("数据必须包含'发出/接收'和'会话时间'列")

    # 转换会话时间列为datetime类型
    df['datetime'] = pd.to_datetime(df['会话时间'], format="%Y/%m/%d %H:%M", errors='coerce')

    # 解析数据
    results = []
    i = 0

    while i < len(df):
        row = df.iloc[i]

        # 如果当前行是"接收"
        if row["发出/接收"] == "接收":
            receive_index = i
            receive_time = row['datetime']

            # 从当前接收位置向后寻找下一个"发出"
            j = i + 1
            while j < len(df):
                if df.iloc[j]["发出/接收"] == "发出":
                    # 找到了"发出"
                    send_index = j
                    send_time = df.iloc[j]['datetime']

                    # 计算时间差（分钟）
                    time_diff = (send_time - receive_time).total_seconds() / 60

                    results.append({
                        'receive_index': receive_index,
                        'send_index': send_index,
                        'receive_time': row['会话时间'],
                        'send_time': df.iloc[j]['会话时间'],
                        'time_diff_minutes': time_diff
                    })

                    # 移动到发出位置之后继续
                    i = j
                    break
                j += 1

            # 如果没有找到对应的"发出"，跳到下一行
            if j == len(df):
                i += 1
        else:
            i += 1

    return results


def process_excel_file(file_content):
    """处理Excel文件内容并返回结果工作簿"""
    # 读取Excel文件
    workbook = load_workbook(io.BytesIO(file_content))
    sheet = workbook.active

    # 找到列名所在的行
    header_row = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and ("发出/接收" in row or "会话时间" in row):
            header_row = i
            break

    if header_row is None:
        raise ValueError("无法找到包含'发出/接收'或'会话时间'的列")

    # 获取列名
    headers = []
    for cell in sheet[header_row]:
        headers.append(cell.value)

    # 找到"发出/接收"和"会话时间"列的索引
    status_col_idx = None
    time_col_idx = None

    for i, header in enumerate(headers):
        if header == "发出/接收":
            status_col_idx = i
        elif header == "会话时间":
            time_col_idx = i

    if status_col_idx is None or time_col_idx is None:
        raise ValueError("必须包含'发出/接收'和'会话时间'列")

    # 读取数据到DataFrame
    data = []
    for i, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not row or row[status_col_idx] is None or row[time_col_idx] is None:
            continue

        data_row = {}
        for j, header in enumerate(headers):
            if j < len(row):
                data_row[header] = row[j]
            else:
                data_row[header] = None

        data.append(data_row)

    df = pd.DataFrame(data)

    # 计算时间差
    try:
        results = calculate_time_differences(df)
    except Exception as e:
        raise ValueError(f"计算时间差时出错: {str(e)}")

    # 添加或找到时间差列
    time_diff_col = None
    for i, header in enumerate(headers):
        if header == "时间差(分钟)":
            time_diff_col = i
            break

    if time_diff_col is None:
        # 添加时间差列
        time_diff_col = len(headers)
        headers.append("时间差(分钟)")
        sheet.cell(row=header_row, column=time_diff_col + 1).value = "时间差(分钟)"
        sheet.cell(row=header_row, column=time_diff_col + 1).font = Font(bold=True)

    # 将结果写入时间差列
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for result in results:
        # 获取接收行的时间差列单元格位置
        cell = sheet.cell(row=result['receive_index'] + header_row + 1, column=time_diff_col + 1)
        cell.value = round(result['time_diff_minutes'], 2)
        cell.fill = highlight_fill
        cell.alignment = Alignment(horizontal='right')

    # 设置列宽
    for i, column in enumerate(sheet.columns, start=1):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(i)].width = adjusted_width

    # 创建汇总表
    if "时间差汇总" in workbook.sheetnames:
        workbook.remove(workbook["时间差汇总"])

    summary_sheet = workbook.create_sheet(title="时间差汇总")
    summary_headers = ["序号", "接收行号", "接收时间", "发出行号", "发出时间", "时间差(分钟)"]

    for i, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=1, column=i)
        cell.value = header
        cell.font = Font(bold=True)

    # 填充汇总数据
    for i, result in enumerate(results, start=1):
        summary_sheet.cell(row=i + 1, column=1).value = i
        summary_sheet.cell(row=i + 1, column=2).value = result['receive_index'] + header_row + 1
        summary_sheet.cell(row=i + 1, column=3).value = result['receive_time']
        summary_sheet.cell(row=i + 1, column=4).value = result['send_index'] + header_row + 1
        summary_sheet.cell(row=i + 1, column=5).value = result['send_time']
        summary_sheet.cell(row=i + 1, column=6).value = round(result['time_diff_minutes'], 2)

    # 设置汇总表列宽
    for i, column in enumerate(summary_sheet.columns, start=1):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        summary_sheet.column_dimensions[get_column_letter(i)].width = adjusted_width

    return workbook


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # 检查是否有文件上传
        if 'file' not in request.files:
            return render_template('index.html', error='没有上传文件')

        file = request.files['file']

        # 如果用户没有选择文件
        if file.filename == '':
            return render_template('index.html', error='没有选择文件')

        # 检查文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            return render_template('index.html', error='请上传Excel文件(.xlsx或.xls)')

        # 读取文件内容
        if file:
            file_content = file.read()

            # 处理Excel文件
            try:
                result_workbook = process_excel_file(file_content)

                # 将处理后的Excel保存到内存
                output = io.BytesIO()
                result_workbook.save(output)
                output.seek(0)

                # 返回Excel文件供下载
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=f"处理结果_{file.filename}",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                return render_template('index.html', error=f'处理文件时出错: {str(e)}')

    return render_template('index.html')


# 创建模板目录
if not os.path.exists('templates'):
    os.makedirs('templates')

# 创建模板文件
with open('templates/index.html', 'w', encoding='utf-8') as f:
    f.write('''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>客服会话时间差计算工具</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { padding: 20px; }
        .container { max-width: 800px; }
        .feature-box { margin-bottom: 20px; }
    </style>
</head>
<body>
    <div class="container">
        <h1 class="mb-4 text-center">客服会话时间差计算工具</h1>

        <div class="card mb-4">
            <div class="card-header bg-primary text-white">
                上传Excel文件
            </div>
            <div class="card-body">
                <form method="post" enctype="multipart/form-data">
                    <div class="mb-3">
                        <label for="file" class="form-label">选择客服会话Excel文件:</label>
                        <input type="file" class="form-control" id="file" name="file" accept=".xlsx,.xls">
                        <div class="form-text">文件应包含'用户昵称', '客服角色', '客服马甲', '会话内容', '发出/接收', '会话时间', '交替标记'等列</div>
                    </div>
                    {% if error %}
                        <div class="alert alert-danger">{{ error }}</div>
                    {% endif %}
                    <button type="submit" class="btn btn-primary w-100">处理文件</button>
                </form>
            </div>
        </div>

        <div class="row">
            <div class="col-md-6">
                <div class="feature-box">
                    <h5>功能说明:</h5>
                    <ul>
                        <li>计算从"接收"到下一个"发出"之间的时间差</li>
                        <li>将结果写入表格的"时间差(分钟)"列</li>
                        <li>生成时间差汇总表</li>
                    </ul>
                </div>
            </div>
            <div class="col-md-6">
                <div class="feature-box">
                    <h5>数据格式要求:</h5>
                    <ul>
                        <li>必须包含"发出/接收"列：标记消息类型</li>
                        <li>必须包含"会话时间"列：格式为 YYYY/MM/DD HH:MM</li>
                        <li>支持标准客服会话表格格式</li>
                    </ul>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
    ''')

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
